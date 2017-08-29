from openpyxl import Workbook
import numpy as np
import pandas as pd
import logging, time

class Excel(object):
	def __init__(self, xls, filename, selected_sheets, 
		Dpname, Dsname, Dgename, Drename, Dorf, Dseq,
		Genome_edit_, Repeat_edit_, ORF_edit_, NCR_edit_):
		"""
		#########################################################################
		#																		#
		#		Made by IL Seob Lee, Software Engineering Department, 2017		#
		#																		#
		#########################################################################
		This program made by Python, PyQt5, Pyinstaller, openpyxl, pandas, numpy.
		If you want to make excutable files(.exe) using Pyinstaller, you must install above modules.

			limits 								# sum limits
			xls 								# Activated xlsx
			sheet_names  						# selected sheets from GUI
			Dp_name								# Duma position col name
			Ds_name 							# Duma sequence col name
			Dge_name							# Genome Structure col name
			Dre_name 							# Repeat Region col name
			Dorf  								# ORF col name
			Dseq 								# Sequence Col name

			length 								# Sheet's length
			P_Raw 								# Removed only '-' data in Sequence

		"""
		self.filename = filename.split('.')[0]
		self.limits = 35.0					
		self.xls = xls 						
		self.sheet_names = selected_sheets 	
		self.Dp_name = Dpname 				
		self.Ds_name = Dsname 				
		self.Dge_name = Dgename 			
		self.Dre_name = Drename 			
		self.Dorf = Dorf 					
		self.Dseq = Dseq					
		self.length = len(self.sheet_names)	
		self.P_Raw = list()
		self.Pxbp = list()
		self.PxLength = list()
		self.PxSum = list()
		self.PxMajor = list()
		self.PxMinor = list()
		self.PxMAF = list()

		self.Genome_idx = self.type_check(Genome_edit_)
		self.Region_idx = self.type_check(Repeat_edit_)
		self.ORF = self.type_check(ORF_edit_)
		self.NCR = self.type_check(NCR_edit_)

		self.P0 = None
		self.P_Raw35 = list()
        # Full Raw Sequence
		for x in self.sheet_names:
			# Raw Data
			tmp = self.xls.parse(x)
			
			if self.P0 is None:
				self.P0 = tmp
			else:
				self.P0 = pd.merge(self.P0, tmp, how="outer", on=[self.Dge_name, self.Dre_name, self.Dorf, self.Dp_name, self.Ds_name])
			self.P0 = self.P0[ self.P0[ self.Ds_name ] != '-' ]
			tmp = tmp[ tmp[ self.Dseq ] != '-']

			self.P_Raw.append(tmp)

			# Raw Length
			self.PxLength.append(len(tmp))
		# backup Original datas
		self.RP_Raw = self.P_Raw.copy()
		self.Rsheets_name = self.sheet_names.copy()

	def init_full(self):
		for bp in self.P_Raw:
			tmp = bp[["A","G","C","T"]]
			psum = tmp.sum(axis=1)
			bp["sum"] = psum
			# 35bp
			tmp2 = bp[[self.Dge_name, self.Dre_name, self.Dorf, "sum"]]
			tmp2 = tmp2[psum >= self.limits]
			self.P_Raw35.append(tmp2)

			tmp = tmp[psum >= self.limits]
			self.Pxbp.append(tmp)
			
			# get Sum
			psum = pd.DataFrame(psum[psum >= self.limits])
			self.PxSum.append(psum)
			
			pmajor, pminor = self.get_major_minor(tmp)

			self.PxMajor.append(pmajor)
			self.PxMinor.append(pminor)

			s, l = self.get_Number_of_GPS(pminor, psum, 5.0, 51.0)
			if l == 0:
				self.PxMAF.append(0)
			else:
				self.PxMAF.append(s/l)

		for bp in self.Pxbp:
			tmp = -bp
			bps = tmp.values.argsort(axis=1)
			bp['minor_idx'] = pd.DataFrame(bps[:,1], index=[bp.index])
			bp['major_idx'] = pd.DataFrame(bps[:,0], index=[bp.index])

	def init_Minor(self, type_):
		PP_Raw = list()
		PP_Length = list()
		PP_sheets = list()
		self.P_Raw35.clear()
		self.Pxbp.clear()
		self.PxSum.clear()
		self.PxMajor.clear()
		self.PxMinor.clear()
		self.PxMAF.clear()
		for i, x in enumerate(self.RP_Raw):
			for j, y in enumerate(self.RP_Raw):
				if i >= j:
					continue
				PP_sheets.append(self.Rsheets_name[i]+" -> "+self.Rsheets_name[j])
				tmp, pmajor, pminor = self.Extract_difference_of_minor(x,y,type_)
				PPxbp = tmp[["A","G","C","T"]]
				psum = PPxbp.sum(axis=1)
				tmp["sum"] = psum

				PP_Raw.append(tmp)
				PP_Length.append(len(tmp))

				tmp2 = tmp[[self.Dge_name, self.Dre_name, self.Dorf, "sum"]]
				tmp2 = tmp2[psum >= self.limits]
				self.P_Raw35.append(tmp2)

				tmp = tmp[psum >= self.limits]
				self.Pxbp.append(tmp[["A","G","C","T"]])
				psum = pd.DataFrame(psum[psum >= self.limits])
				psum = psum.loc[tmp.index]
				pmajor = pmajor.loc[tmp.index]
				pminor = pminor.loc[tmp.index]
				self.PxSum.append(psum)
				self.PxMajor.append(pmajor)
				self.PxMinor.append(pminor)
				
				s, l = self.get_Number_of_GPS(pminor,psum, 5.0, 51.0) if len(tmp)!=0 else (0,0)
				if l == 0:
					self.PxMAF.append(0)
				else:
					self.PxMAF.append(s/l)

		self.P_Raw = PP_Raw
		self.PxLength = PP_Length
		self.sheet_names = PP_sheets
		self.length = len(self.sheet_names)

		for bp in self.Pxbp:
			tmp = -bp
			bps = tmp.values.argsort(axis=1)
			bp['minor_idx'] = pd.DataFrame(bps[:,1], index=[bp.index]) if len(bp)!=0 else 0
			bp['major_idx'] = pd.DataFrame(bps[:,0], index=[bp.index]) if len(bp)!=0 else 0

	def Extract_difference_of_minor(self, x, y, type_):
		if len(x) > len(y):
			y = pd.DataFrame(y, index = x.index)
		else:
			x = pd.DataFrame(x, index = y.index)
        
		tmp_x = x[ x[self.Dp_name] == y[self.Dp_name]]
		tmp_y = y[ x[self.Dp_name] == y[self.Dp_name]]
		x_bp = -tmp_x[["A","G","C","T"]]
		x_bps = x_bp.values.argsort(axis=1)
		x_minor_idx = pd.DataFrame(x_bps[:,1], index=[x_bp.index])
		x_major_idx = pd.DataFrame(x_bps[:,0], index=[x_bp.index])
    
		y_bp = -tmp_y[["A","G","C","T"]]
		y_bps = y_bp.values.argsort(axis=1)
		y_minor_idx = pd.DataFrame(y_bps[:,1], index=[y_bp.index])
		y_major_idx = pd.DataFrame(y_bps[:,0], index=[y_bp.index])
    
		tmp_x = tmp_x[ (x_major_idx == y_major_idx).values.tolist() ]
		tmp_y = tmp_y[ (x_major_idx == y_major_idx).values.tolist() ]
    
		x_major, x_minor = self.get_major_minor(tmp_x[["A","G","C","T"]])
		y_major, y_minor = self.get_major_minor(tmp_y[["A","G","C","T"]])
		x_minor = pd.DataFrame(x_minor.values, index = tmp_x.index)
		y_minor = pd.DataFrame(y_minor.values, index = tmp_y.index)

		x_sum = pd.DataFrame(tmp_x[["A","G","C","T"]].sum(axis=1), index = tmp_x.index)
		y_sum = pd.DataFrame(tmp_y[["A","G","C","T"]].sum(axis=1), index = tmp_y.index)
    
		x_maf = np.divide(x_minor, x_sum) * 100
		y_maf = np.divide(y_minor, y_sum) * 100
		if type_ == 'ASC':
			return tmp_x[((y_maf - x_maf) >= 5.0).values] , x_major.loc[tmp_x.index], x_minor.loc[tmp_x.index]
		else:
			return tmp_x[((x_maf - y_maf) >= 5.0).values] , x_major.loc[tmp_x.index], x_minor.loc[tmp_x.index]

	def type_check(self, L):
		for i in range(0,len(L)):
			try:
				L[i] = int(L[i])
			except:
				L[i] = str(L[i])
		return L

	def Analyze(self, path,Analyze_type):
		try:
			self.Analyze_type = Analyze_type
			if Analyze_type == "Difference_of_Minor":
				self.s1 = [ 0, 2.5, 5, 15, 25, 5]
				self.s2 = [ 2.5, 5, 15, 25, 51.0, 51.0]
				types_ = ["ASC", "DESC"]
				for i in range(0,2):
					wb = Workbook()
					ws1 = wb.active
					ws2 = wb.create_sheet("GPS")
					ws3 = wb.create_sheet("Genome structure")
					ws4 = wb.create_sheet("ORF")
					ws5 = wb.create_sheet("NCR")
					ws6 = wb.create_sheet("Base composition_1")
					ws7 = wb.create_sheet("Base composition_2")

					self.init_Minor(types_[i])
					self.sheet1_m(ws1, types_[i])
					self.sheet2(ws2)			
					self.sheet3(ws3)			
					self.sheet4_5(ws4, "ORF")			
					self.sheet4_5(ws5, "NCR")			
					self.sheet6(ws6)			
					self.sheet7(ws7)
					wb.save(path+"/["+types_[i]+"분석]" + self.filename + '.xlsx')
			else:
				self.s1 = [ 2.5, 5, 15, 25, 5]
				self.s2 = [ 5, 15, 25, 51.0, 51.0]
				wb = Workbook()
				ws1 = wb.active
				ws2 = wb.create_sheet("GPS")
				ws3 = wb.create_sheet("Genome structure")
				ws4 = wb.create_sheet("ORF")
				ws5 = wb.create_sheet("NCR")
				ws6 = wb.create_sheet("Base composition_1")
				ws7 = wb.create_sheet("Base composition_2")
				self.init_full()
				self.sheet1(ws1)
				
				self.sheet2(ws2)			
				self.sheet3(ws3)			
				self.sheet4_5(ws4, "ORF")			
				self.sheet4_5(ws5, "NCR")			
				self.sheet6(ws6)			
				self.sheet7(ws7)
				wb.save(path+"/[분석]" + self.filename + '.xlsx')
			self.P0.to_excel(path+'/[통합]'+self.filename+'.xlsx', index=False)
			return "Success"
		except PermissionError:
			return "Permission"
		except Exception as e:
			logging.error("{0} {1}".format(time.time(), e))
			return "Error"

	def get_major_minor(self,Passage):
		col = ["A","G","C","T"]
		ranked_df = Passage.apply(np.argsort, axis=1)
		data = np.array(list(ranked_df.C.values))
		arr = np.zeros((len(Passage),1))
		for i, d in enumerate(data):
			arr[i] = Passage[i:i+1][col[d]]
		minor = pd.DataFrame(arr)
		major = pd.DataFrame(Passage.max(axis=1).reset_index()[0])
		return major, minor

	def get_Number_of_GPS(self, minor_, sum_, s1, s2):
		maf_ = np.divide(minor_, sum_) * 100
		idx = np.logical_and(maf_>=s1, maf_<s2)
		idx = idx.values.tolist()
		return maf_[idx].sum()[0], len(minor_[idx])

	def get_level_major_minor(self, major, minor, sum_, s1, s2):
		maf_ = np.divide(minor, sum_) * 100
		idx = np.logical_and(maf_ >= s1, maf_ < s2)
		idx = idx.values.tolist()
		return pd.DataFrame(major[idx].values, index=[sum_[idx].index]), pd.DataFrame(minor[idx].values, index=[sum_[idx].index])
		
	# genome, minor, sum 병합 함수
	def merge_Genome_Structure(self, Passage, minor_, sum_, title):
		df_gn = Passage[[title]]
		df_gn = df_gn[ Passage["sum"] >= 35.0 ]
		df_gn_minor = pd.concat([pd.DataFrame(df_gn.values), pd.DataFrame(minor_.values), pd.DataFrame(sum_.values)], axis=1)
		return pd.DataFrame(df_gn_minor.values, columns=[title, "minor", "sum"], index=[sum_.index])
		
	def insert_value_in_cell(self,ws, rows, cols, Passage, minor_, sum_, s1, s2, title, types):
		"""
			# rows = [4,5,6,7,8,9]
			# cols = "D"
			# index = ["TRL", "UL", "IRL", "IRS", "US", "TRS"]
			# title = self.Dge_name : "Genome\nstrucure",  self.Dre_name  : "Repeat\nregion"
			# type = "GPS", "MAF"

			# 3개 컬럼에서 index별로 자르고
			# 갯수를 각 셀에 삽입
			# 

		"""
		if title == self.Dge_name:
			index = self.Genome_idx
		else:
			index = self.Region_idx

		# For Minor Exception
		if len(Passage) == 0:
			for i in range(0, len(rows)):
				ws[cols+str(rows[i])] = 0 if types == "GPS" else '0.000%'
			ws[cols+str(rows[len(rows)-1] + 1)] = 0 if types == "GPS" else '0.000%'
		else:
			merged = self.merge_Genome_Structure(Passage, minor_, sum_, title)
			_, tm_minor = self.get_level_major_minor(minor_, minor_, sum_, s1,s2)
			mid_rows = merged.loc[ tm_minor.index ]

			cnt_sum=0;
			for i in range(0, len(rows)): # rows[i], index[i]		
				# 각 index 별로 그룹핑
				mrows = mid_rows [ mid_rows[ title ] == index[i] ]
				if types == "GPS":
					ws[cols+str(rows[i])] = len(mrows)
					cnt_sum+=len(mrows)
				elif types == "MAF":
					if len(mrows) != 0:
						maf_ = np.divide(mrows["minor"].to_frame(), mrows["sum"].to_frame()) * 100
						ws[cols+str(rows[i])] = str(round(maf_.sum()[0] / len(maf_),3)) + '%'
					else:
						ws[cols+str(rows[i])] = '0.000%'
			if types == "GPS":
				ws[cols+str(rows[len(rows)-1] + 1)] = cnt_sum
			elif types == "MAF":
				s, l = self.get_Number_of_GPS(minor_, sum_, s1, s2)
				if l == 0:
					ws[cols+str(rows[len(rows)-1] + 1)] = '0.000%'
				else:	
					ws[cols+str(rows[len(rows)-1] + 1)] = str(round(s / l, 3)) + '%'

	def next_col(self, asc):
		asc2 = list(asc)
		if len(asc2) == 1 and asc2[0] != 'Z':
			return chr(ord(asc2[0])+1)
		elif len(asc2) == 1 and asc2[0] == 'Z':
			return 'AA'
		else:
			if asc[-1] != 'Z':
				return asc[:-1] + chr(ord(asc[-1])+1)
			else:
				p = None
				for i in range(-1, -len(asc)-1, -1):
					p = i
					if asc2[i] == 'Z':
						asc2[i] = 'A'
					else:
 						break
				if p == -len(asc) and asc[p] == 'Z':
					return 'A'+''.join(asc2)
				else:
					asc = ''.join(asc2)
					return asc[:p] + chr(ord(asc[p])+1) + asc[p+1:]

	def sheet1_m(self, ws, types_):
		logging.info("{0} Start Sheet1_m".format(time.time()))
		ws.title = "Polymorphic site"
		# A col
		ws['A1'] = types_
		ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
		
		ws['A3'] = self.filename
		ws.merge_cells(start_row=3, start_column=1, end_row=2+self.length, end_column=1)
		# C col
		ws['C1'] = "Genome length (bp)"
		ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
		# D col
		ws['D1'] = "Genome length (35이상)"
		ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
		
		ws['E1'] = "Average of MAF"
		ws.merge_cells(start_row=1, start_column=5, end_row=2, end_column=5)

		ws['F1'] = "Number of polymorphic sites"
		ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=11)

		ws['F2'] = "0≤n<2.5"
		ws['G2'] = "2.5≤n<5"
		ws['H2'] = "5≤n<15"
		ws['I2'] = "15≤n<25"
		ws['J2'] = "25≤n"
		ws['K2'] = "5≤n"
		for i in range(self.length):
			ws['B' + str(i+3)] = self.sheet_names[i]
			ws['C' + str(i+3)] = self.PxLength[i]
			ws['D' + str(i+3)] = len(self.P_Raw35[i])
			ws['E' + str(i+3)] = str(round(self.PxMAF[i], 3)) + '%'
			ws['F' + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 0.0, 2.5)[1] if len(self.PxMinor[i])!=0 else 0
			ws['G' + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 2.5, 5.0)[1] if len(self.PxMinor[i])!=0 else 0
			ws["H" + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 5.0, 15.0)[1] if len(self.PxMinor[i])!=0 else 0
			ws["I" + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 15.0, 25.0)[1] if len(self.PxMinor[i])!=0 else 0
			ws["J" + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 25.0, 51.0)[1] if len(self.PxMinor[i])!=0 else 0
			ws["K" + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 5.0, 51.0)[1] if len(self.PxMinor[i])!=0 else 0
		logging.info("{0} End Sheet1_m".format(time.time()))

	def sheet1(self,ws):
		logging.info("{0} Start Sheet1".format(time.time()))
		ws.title = "Polymorphic site"
		ws['A1'] = "strain"
		ws['D1'] = "Average of MAF"
		ws['E1'] = "Number of polymorphic sites"
		ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=10)
		ws['E2'] = "2.5≤n<5"
		ws['F2'] = "5≤n<15"
		ws['G2'] = "15≤n<25"
		ws['H2'] = "25≤n"
		ws['I2'] = "5≤n"
		# A col
		ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
		ws['C1'] = "Genome length (bp)"
		ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)		
		# D col
		ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
		ws['A3'] = self.filename
		ws.merge_cells(start_row=3, start_column=1, end_row=2+self.length, end_column=1)
		for i in range(self.length):
			ws['B' + str(i+3)] = self.sheet_names[i]
			ws['C' + str(i+3)] = self.PxLength[i]
			ws['D' + str(i+3)] = str(round(self.PxMAF[i], 3)) + '%'
			ws['E' + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 2.5, 5.0)[1]
			ws["F" + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 5.0, 15.0)[1]
			ws["G" + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 15.0, 25.0)[1]
			ws["H" + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 25.0, 51.0)[1]
			ws["I" + str(i+3)] = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 5.0, 51.0)[1]
		logging.info("{0} End Sheet1".format(time.time()))

	def sheet2(self,ws):
		logging.info("{0} Start Sheet2".format(time.time()))
		ws['A1'] = "strain"
		ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
		ws['C1'] = "Genome length (bp)"
		ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
		ws['D1'] = "Average of MAF at GPS"
		ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=9)
		ws['D2'] = "2.5≤n<5"
		ws['E2'] = "5≤n<15"
		ws['F2'] = "15≤n<25"
		ws['G2'] = "25≤n"
		ws['H2'] = "5≤n"

		ws['A3'] = self.filename
		ws.merge_cells(start_row=3, start_column=1, end_row=2+self.length, end_column=1)
		for i in range(self.length):
			ws['B' + str(i+3)] = self.sheet_names[i]
			ws['C' + str(i+3)] = self.PxLength[i]

			s, l = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 2.5, 5.0) if len(self.PxMinor[i])!=0 else (0,0)
			if l == 0:
				ws['D' + str(i+3)] = '0.000%'
			else:
				ws['D' + str(i+3)] = str(round(s/l,3)) + '%'

			s, l = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 5.0, 15.0) if len(self.PxMinor[i])!=0 else (0,0)
			if l == 0:
				ws['E' + str(i+3)] = '0.000%'
			else:	
				ws['E' + str(i+3)] = str(round(s/l,3)) + '%'

			s, l = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 15.0, 25.0) if len(self.PxMinor[i])!=0 else (0,0)
			if l == 0:
				ws['F' + str(i+3)] = '0.000%'
			else:
				ws['F' + str(i+3)] = str(round(s/l,3)) + '%'

			s, l = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 25.0, 51.0) if len(self.PxMinor[i])!=0 else (0,0)
			if l == 0:
				ws['G' + str(i+3)] = '0.000%'
			else:
				ws['G' + str(i+3)] = str(round(s/l,3)) + '%'

			s, l = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 5.0, 51.0) if len(self.PxMinor[i])!=0 else (0,0)
			if l==0:
				ws['H' + str(i+3)] = '0.000%'
			else:
				ws['H' + str(i+3)] = str(round(s/l,3)) + '%'

			if self.Analyze_type == "Difference_of_Minor":
				ws['I2'] = "0≤n<2.5"
				s, l = self.get_Number_of_GPS(self.PxMinor[i], self.PxSum[i], 0, 2.5) if len(self.PxMinor[i])!=0 else (0,0)
				if l==0:
					ws['I' + str(i+3)] = '0.000%'
				else:
					ws['I' + str(i+3)] = str(round(s/l,3)) + '%'
		logging.info("{0} End Sheet2".format(time.time()))

	def Full_Seq_in_sheet3(self, ws, col):
		logging.info("{0} Start Full_Seq_in_sheet3".format(time.time()))
		for r in range(4, 4+len(self.Genome_idx)):
			ws[col+str(r)] = self.Genome_idx[r-4]
		ws[col+str(4+len(self.Genome_idx))] = 'Total'
		for r in range(0, len(self.Region_idx)):
			row = r+5+len(self.Genome_idx)
			ws[col+str(row)] = self.Region_idx[r]
		ws[col+str(5+len(self.Genome_idx)+len(self.Region_idx))] = 'Total'

		rs_ = 5+len(self.Genome_idx)+len(self.Region_idx)
		ws[col+str(rs_+1)] = 'NCR'
		ws[col+str(rs_+2)] = 'ORF'

		col = self.next_col(col)
		for i in range(0, self.length):
			ws[col+str(2)] = self.sheet_names[i]

			# 35 이하 포함
			ws[col+str(3)] = '(bp_full)Length'
			sum_ = 0
			for r in range(4, 4+len(self.Genome_idx)):
				tmp = len(self.P_Raw[i][ self.P_Raw[i][ self.Dge_name] == self.Genome_idx[r-4]]) if len(self.P_Raw[i])!=0 else 0
				ws[col+str(r)] = tmp
				sum_ += tmp
			ws[col+str(4+len(self.Genome_idx))] = sum_
			sum_=0
			for r in range(0, len(self.Region_idx)):
				row = r+5+len(self.Genome_idx)
				tmp = len(self.P_Raw[i][ self.P_Raw[i][ self.Dre_name] == self.Region_idx[r]]) if len(self.P_Raw[i])!=0 else 0
				ws[col+str(row)] = tmp
				sum_ += tmp
			ws[col+str(5+len(self.Genome_idx)+len(self.Region_idx))] = sum_

			# ORF & NCR
			ws[col+str(rs_+1)] = len(self.P_Raw[i][ self.P_Raw[i][ self.Dorf].isin(self.ORF)]) if len(self.P_Raw[i])!=0 else 0
			ws[col+str(rs_+2)] = len(self.P_Raw[i][ self.P_Raw[i][ self.Dorf].isin(self.NCR)]) if len(self.P_Raw[i])!=0 else 0

			# 35 이하 포함 X = 35 이상
			col = self.next_col(col)
			ws[col+str(3)] = '(bp_35이상)Length'
			sum_ = 0
			for r in range(4, 4+len(self.Genome_idx)):
				tmp = len(self.P_Raw35[i][ self.P_Raw35[i][ self.Dge_name] == self.Genome_idx[r-4]]) if len(self.P_Raw[i])!=0 else 0
				ws[col+str(r)] = tmp
				sum_ += tmp
			ws[col+str(4+len(self.Genome_idx))] = sum_
			sum_=0
			for r in range(0, len(self.Region_idx)):
				row = r+5+len(self.Genome_idx)
				tmp = len(self.P_Raw35[i][ self.P_Raw35[i][ self.Dre_name] == self.Region_idx[r]]) if len(self.P_Raw35[i])!=0 else 0
				ws[col+str(row)] = tmp
				sum_ += tmp
			ws[col+str(5+len(self.Genome_idx)+len(self.Region_idx))] = sum_

			ws[col+str(rs_+1)] = len(self.P_Raw35[i][ self.P_Raw35[i][ self.Dorf].isin(self.ORF)]) if len(self.P_Raw35[i])!=0 else 0
			ws[col+str(rs_+2)] = len(self.P_Raw35[i][ self.P_Raw35[i][ self.Dorf].isin(self.NCR)]) if len(self.P_Raw35[i])!=0 else 0
			col = self.next_col(col)
		logging.info("{0} End Full_Seq_in_sheet3".format(time.time()))
	# s1, s2 비율, Genome, Region의 길이에 따른 row 길이 변경...
	def sheet3(self,ws):
		logging.info("{0} Start sheet3".format(time.time()))
		col = 'D'
		for i in range(0, len(self.s1)):
			r = i*(10+len(self.Genome_idx) + len(self.Region_idx)+5)+ 1
			ws['A'+str(r)] = "Region"
			ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=2)
			ws['C'+str(r)] = "Length (bp)"
			ws.merge_cells(start_row=r, start_column=3, end_row=r+2, end_column=3)

			ws['A'+str(r+3)] = "Genome Structure"
			ws.merge_cells(start_row=r+3, start_column=1, end_row=r+3+len(self.Genome_idx), end_column=1)
			ws['A'+str(r+10)] = "Repeat region"
			ws.merge_cells(start_row=r+4+len(self.Genome_idx), start_column=1, end_row=r+4+len(self.Genome_idx)+len(self.Region_idx), end_column=1)
			
			rows = list(range(r+3, r+3+len(self.Genome_idx)))
			cnt_num = 0
			for ix in range(0, len(self.Genome_idx)):
				ws['B' + str(rows[ix])] = self.Genome_idx[ix]
				nlen = len(self.P0[ self.P0[ self.Dge_name ] == self.Genome_idx[ix]])
				ws['C' + str(rows[ix])] = nlen
				cnt_num += nlen
			ws['B' + str(r+3+len(self.Genome_idx))] = "Total"
			ws['C' + str(r+3+len(self.Genome_idx))] = cnt_num
			cnt_num = 0

			nr = r+4+len(self.Genome_idx)+len(self.Region_idx)

			nrows = list(range(r+4+len(self.Genome_idx), nr))
			for ix in range(0, len(self.Region_idx)):
				ws['B' + str(nrows[ix])] = self.Region_idx[ix]
				nlen = len(self.P0[ self.P0[ self.Dre_name ] == self.Region_idx[ix]])
				ws['C'+str(nrows[ix])] = nlen
				cnt_num += nlen
			ws['B'+str(nr)] = 'Total'
			ws['C' + str(nr)] = cnt_num

			## ORF NCR
			ws["A"+str(nr+1)] = "ORF"
			ws["A"+str(nr+2)] = "NCR"
			# ORF NCR Full Length
			ws['C'+str(nr+1)] = len(self.P0[ self.P0[self.Dorf].isin(self.ORF)])
			ws['C'+str(nr+2)] = len(self.P0[ self.P0[self.Dorf].isin(self.NCR)])

			col = 'D'
			ncol = 'D'
			for x in range(0,3):
				for c in range(0, self.length):
					ws[col+str(r+2)] = self.sheet_names[c]
					if x==0 or x==1:
						# 35 이상
						px_orf_merged = self.merge_Genome_Structure(self.P_Raw[c], self.PxMinor[c],
										self.PxSum[c], self.Dorf)
						_, px_minor = self.get_level_major_minor(self.PxMinor[c], self.PxMinor[c],
										self.PxSum[c], self.s1[i], self.s2[i])
						tx_rows = px_orf_merged.loc[ px_minor.index ]
						tx_orf = tx_rows[ tx_rows[ self.Dorf ].isin(self.ORF)]
						tx_ncr = tx_rows[ tx_rows[ self.Dorf ].isin(self.NCR)]

						if x==0:
							ws[col+str(r)] = "Number of GPS"
							ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=4 + self.length - 1)
							ws[col+str(r+1)] = self.filename
							ws.merge_cells(start_row=r+1, start_column=4, end_row=r+1, end_column=4 + self.length - 1)

							self.insert_value_in_cell(ws, rows, col, self.P_Raw[c], self.PxMinor[c], 
										self.PxSum[c] , self.s1[i], self.s2[i],  self.Dge_name , "GPS")
							self.insert_value_in_cell(ws, nrows, col, self.P_Raw[c], self.PxMinor[c], 
										self.PxSum[c] , self.s1[i], self.s2[i],  self.Dre_name , "GPS")
							ws[col+str(nr+1)] = len(tx_orf)
							ws[col+str(nr+2)] = len(tx_ncr)
						elif x==1:
							ws[col+str(r)] = "Average MAF at GPS"
							ws.merge_cells(start_row=r, start_column=4 + self.length, end_row=r, end_column=4 + 2 *self.length - 1)
							ws[col+str(r+1)] = self.filename
							ws.merge_cells(start_row=r+1, start_column=4 + self.length, end_row=r+1, end_column=4 + 2 *self.length - 1)

							self.insert_value_in_cell(ws, rows, col, self.P_Raw[c], self.PxMinor[c], 
										self.PxSum[c] , self.s1[i], self.s2[i],  self.Dge_name , "MAF")
							self.insert_value_in_cell(ws, nrows, col, self.P_Raw[c], self.PxMinor[c], 
										self.PxSum[c] , self.s1[i], self.s2[i],  self.Dre_name , "MAF")

							if len(tx_orf) == 0:
								ws[col+str(nr+1)] = '0.000%'
							else:	
								ws[col+str(nr+1)] = str(round((np.divide(tx_orf[["minor"]], self.PxSum[c].loc[ tx_orf.index]) * 100).sum()[0]/len(tx_orf), 3)) +'%'

							if len(tx_ncr) == 0:
								ws[col+str(nr+2)] = '0.000%'
							else:	
								ws[col+str(nr+2)] = str(round((np.divide(tx_ncr[["minor"]], self.PxSum[c].loc[ tx_ncr.index]) * 100).sum()[0]/len(tx_ncr), 3)) +'%'
					else:
						ws[col+str(r)] = "Number of GPS / length"
						ws.merge_cells(start_row=r, start_column=4 + 2 *self.length, end_row=r, end_column=4 + 3 *self.length - 1)
						ws[col+str(r+1)] = self.filename
						ws.merge_cells(start_row=r+1, start_column=4 + 2 *self.length, end_row=r+1, end_column=4 + 3 *self.length - 1)
						sum_ = 0
						ridx = 0
						for rs in range(r+3, r+5+len(self.Genome_idx)+len(self.Region_idx)):
							if ridx < len(self.Genome_idx):
								llen_ = len(self.P_Raw35[c][ self.P_Raw35[c][ self.Dge_name] == self.Genome_idx[ridx]])
								sum_ += llen_
							elif ridx == len(self.Genome_idx):
								llen_ = sum_
								sum_ = 0
							elif len(self.Genome_idx) < ridx and ridx < len(self.Genome_idx)+1+len(self.Region_idx):
								llen_ = len(self.P_Raw35[c][ self.P_Raw35[c][ self.Dre_name] == self.Region_idx[ridx-1-len(self.Genome_idx)]])
								sum_ += llen_
							else:
								llen_ = sum_

							# ncol = chr(ord(col)-self.length*2)
							if llen_ == 0:
								ws[col + str(rs)] = '0.00000%'
							else:
								ws[col + str(rs)] = ws[ncol+str(rs)].value / llen_

							ridx = ridx + 1

						# ORF & NCR
						for rs in range(2):
							rs_ = r+5+len(self.Genome_idx)+len(self.Region_idx) + rs
							if rs == 0:
								tmp = len(self.P_Raw35[c][ self.P_Raw35[c][self.Dorf].isin(self.ORF)])
								if tmp == 0:
									ws[col+str(rs_)] = '0.00000%'
								else:
									ws[col+str(rs_)] = ws[ncol+str(rs_)].value / tmp
							else:
								tmp = len(self.P_Raw35[c][ self.P_Raw35[c][self.Dorf].isin(self.NCR)])
								if tmp == 0:
									ws[col+str(rs_)] = '0.00000%'
								else:
									ws[col+str(rs_)] = ws[ncol+str(rs_)].value / tmp
						ncol = self.next_col(ncol)

					col = self.next_col(col)

			col = self.next_col(col)
			if self.s2[i] != 51.0:
				ws[col+str(r+3)] = str(self.s1[i]) + "~" + str(self.s2[i]) + "%"
			else:
				ws[col+str(r+3)] = str(self.s1[i]) + "% 이상"

		#### Full Sequence				
		col = self.next_col(col)
		col = self.next_col(col)
		self.Full_Seq_in_sheet3(ws, col)
		logging.info("{0} End sheet3".format(time.time()))

	def sheet4_5(self, ws, title):
		logging.info("{0} Start sheet4_5".format(time.time()))
		ws['B2'] = self.filename
		ws["C2"] = "Length"
		ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=3+self.length-1)
		ws["B3"] = title
		if title == 'ORF':
			orf_ = self.ORF
		else:
			orf_ = self.NCR
		col = 'C'
		rows = list(range(4, len(orf_) + 4))

		for i in range(0, self.length):
			ws[col + str(3)] = self.sheet_names[i]
			col = self.next_col(col)

		# Full - 35이하 포함
		col = 'B'
		for c in range(0, self.length+1):
			cnt_ = 0
			for i in range(0, len(orf_)):
				if c==0:
					ws[col+str(rows[i])] = orf_[i]
				else:
					tmp = len(self.P_Raw[c-1][ self.P_Raw[c-1][self.Dorf] == orf_[i] ])
					ws[col+str(rows[i])] = tmp
					cnt_ += tmp
			if c is 0:
				ws[col + str(4+len(orf_))] = 'total'
			else:
				ws[col + str(4+len(orf_))] = cnt_
			col = self.next_col(col)

		px_orf_merged = list()
		for n in range(0, self.length):
			px_orf_merged.append(self.merge_Genome_Structure(self.P_Raw[n], self.PxMinor[n], self.PxSum[n], self.Dorf))

		col = chr(ord('A')+3+self.length)
		for i in range(0, len(self.s1)):
			co = (2+3*self.length) * i + (5+self.length)
			if self.s2[i] != 51.0:
				ws[col+str(2)] = str(self.s1[i]) + '~' + str(self.s2[i]) +'%'
			else:
				ws[col+str(2)] = str(self.s1[i]) + '%~'
			ws[col+str(3)] = title
			for ix in range(0, len(orf_)):
				ws[col+str(rows[ix])] = orf_[ix]
			ws[col+str(len(orf_) + 4)] = 'total'

			gps = list()
			col = self.next_col(col)			
			ws[col+str(2)] = "Number of GPS"
			ws.merge_cells(start_row=2, start_column=co, end_row=2, end_column=co+self.length-1)
			for n in range(0, self.length):
				ws[col+str(3)] = self.sheet_names[n]
				_, tx_minor = self.get_level_major_minor(self.PxMinor[n], self.PxMinor[n], self.PxSum[n],self.s1[i], self.s2[i])
				tx_rows = px_orf_merged[n].loc[ tx_minor.index ]

				g = list()
				cnt_ = 0
				for ix in range(0, len(orf_)):
					mrows = tx_rows[ tx_rows[self.Dorf] == orf_[ix] ]
					ws[col+str(rows[ix])] = len(mrows)
					cnt_ += len(mrows)
					g.append(len(mrows))
				ws[col+str(len(orf_) + 4)] = cnt_
				gps.append(g)
				col = self.next_col(col)


			# Average MAF at GPS
			ws[col+str(2)] = "Average MAF at GPS"
			ws.merge_cells(start_row=2, start_column=co+self.length, end_row=2, end_column=co+2*self.length-1)
			for n in range(0, self.length):
				ws[col+str(3)] = self.sheet_names[n]
				_, tx_minor = self.get_level_major_minor(self.PxMinor[n], self.PxMinor[n], self.PxSum[n],self.s1[i], self.s2[i])
				tx_rows = px_orf_merged[n].loc[ tx_minor.index ]
				for ix in range(0, len(orf_)):
					mrows = tx_rows[ tx_rows[self.Dorf] == orf_[ix] ]
					if len(mrows) == 0:
						ws[col+str(rows[ix])] = '0.000%'
					else:
						maf_ = np.divide(mrows["minor"].to_frame(), mrows["sum"].to_frame()) * 100
						val = maf_.sum()[0] / len(maf_)
						ws[col+str(rows[ix])] = str(round(val, 3)) + '%'
				col = self.next_col(col)

			ws[col+str(2)] = "Number of GPS / Length"
			ws.merge_cells(start_row=2, start_column=co+2*self.length, end_row=2, end_column=co+3*self.length-1)
			ncol = 'C'
			for n in range(0, self.length):
				ws[col+str(3)] = self.sheet_names[n]
				for ix in range(0, len(orf_)):
					l_ = int(str(ws[ncol+str(rows[ix])].value))
					if l_ == 0:
						ws[col+str(rows[ix])] = '0.00000%'	
					else:
						ws[col+str(rows[ix])] = gps[n][ix] / l_
				ncol = self.next_col(ncol)
				col = self.next_col(col)

			col = self.next_col(col)
		logging.info("{0} End sheet4_5".format(time.time()))

	def sheet6(self,ws):
		logging.info("{0} Start sheet6".format(time.time()))
		col = 'B'
		for i in range(0, len(self.s1)):
			if self.s2[i] == 51.0:
				ws[col+'2'] = str(self.s1[i]) + '%~'
			else:
				ws[col+'2'] = str(self.s1[i]) + '~' + str(self.s2[i]) +'%'

			# get bp data from between s1 and s2
			mxr = list()

			for p in range(0, self.length):
				mj, mn = self.get_level_major_minor(self.PxMajor[p], self.PxMinor[p], self.PxSum[p], self.s1[i], self.s2[i])
				mxr.append(self.Pxbp[p].loc[mj.index])

			kcol = col

			for c in range(-1, self.length):
				if c == -1:
					ws[kcol+'3'] = "Major/Minor"
				else:
					ws[kcol+'3'] = self.sheet_names[c]
				kcol = self.next_col(kcol)

			
			bp_name = ['A', 'G', 'C', 'T']
			for x in range(-1, self.length):
				rows = 4
				for a in range(0, len(bp_name)):
					for b in range(0, len(bp_name)):
						if a==b:
							continue
						if x == -1:
							ws[col+str(rows)] = bp_name[a] + '/' + bp_name[b]
						else:
							ws[col+str(rows)] = len(mxr[x][np.logical_and( mxr[x]['major_idx']==a, mxr[x]['minor_idx']==b )])
						rows = rows + 1
				col = self.next_col(col)
			col = self.next_col(col)
		logging.info("{0} End sheet6".format(time.time()))

	def sheet7(self, ws):
		logging.info("{0} Start sheet7".format(time.time()))
		bp_name = ['A', 'G', 'C', 'T']
		for i in range(0, len(self.s1)):
			r = i * (self.length + 5) + 2

			if self.s2[i] == 51.0:
				ws["A"+str(r)] = str(self.s1[i]) + "%~"
			else:
				ws["A"+str(r)] = str(self.s1[i]) + "~" + str(self.s2[i]) + "%"

			ws["B"+str(r)] = "Virus"
			ws.merge_cells(start_row=r, start_column=2, end_row=r+2, end_column=2)

			ws["C"+str(r)] = "Number of GPS"
			ws.merge_cells(start_row=r, start_column=3, end_row=r+2, end_column=3)

			ws["D"+str(r)] = "GPS Mean"
			ws.merge_cells(start_row=r, start_column=4, end_row=r+2, end_column=4)

			ws["E"+str(r)] = "Mj"
			ws.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=5)
			ws["E"+str(r+2)] = "Mn"

			ws['F'+str(r)] = 'A'
			ws.merge_cells(start_row=r, start_column=6, end_row=r+1, end_column=8)
			ws['I'+str(r)] = 'G'
			ws.merge_cells(start_row=r, start_column=9, end_row=r+1, end_column=11)
			ws['L'+str(r)] = 'C'
			ws.merge_cells(start_row=r, start_column=12, end_row=r+1, end_column=14)
			ws['O'+str(r)] = 'T'
			ws.merge_cells(start_row=r, start_column=15, end_row=r+1, end_column=17)

			for x in range(0, self.length):
				row = r+x+3
				mj, mn = self.get_level_major_minor(self.PxMajor[x], self.PxMinor[x], self.PxSum[x], self.s1[i], self.s2[i])
				mxr = self.Pxbp[x].loc[mj.index]
				s, l = self.get_Number_of_GPS(self.PxMinor[x], self.PxSum[x], self.s1[i], self.s2[i]) if len(self.PxMinor[x])!=0 else (0,0)
				ws["B"+str(row)] = self.sheet_names[x]
				ws['C'+str(row)] = len(mn)
				if l == 0:
					ws['D'+str(row)] = '0.000%'
				else:	
					ws['D'+str(row)] = str(round(s/l, 3)) + '%'
				col = 'F'
				for a in range(0, 4):
					for b in range(0, 4):
						if a == b:
							continue
						else:
							ws[col+str(r+2)] = bp_name[b].lower()
							ws[col+str(row)] = len(mxr[np.logical_and( mxr['major_idx']==a, mxr['minor_idx']==b )])
						col = self.next_col(col)
		logging.info("{0} End sheet7".format(time.time()))
